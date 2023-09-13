import openpyxl
from openpyxl import Workbook, load_workbook
path="test1.xlsx"
wb=openpyxl.load_workbook(path)
total=0                    

ws=wb.active
max_row=ws.max_row
print(max_row)
for i in range(2,max_row+1):
    hours=ws['B'+str(i)].value
    rate=ws['C'+str(i)].value
    if (type(hours)!=str and type(rate)!=str):
        salary=hours*rate
        ws['D'+str(i)].value=salary
        salary1=round(salary, 2)
        print(salary1)
    if salary1>3000:
        total+=1
        
print()
wb.save('result.xlsx')

wb.close()